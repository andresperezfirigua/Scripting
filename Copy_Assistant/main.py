import keyboard
import openpyxl
import pyperclip
import time
import config

# Load the Excel workbook
wb = openpyxl.load_workbook(config.FILE_PATH)
sheet = wb.active

# Get the maximum row number in the column
max_row = sheet.max_row

# Copy the first cell value to the clipboard
cell_value = sheet.cell(row=2, column=2).value
pyperclip.copy(str(cell_value))

# Loop through each row and copy the cell value to the clipboard
for row in range(3, max_row + 1):
    # Wait for the user to paste the current cell value using "CTRL+V"
    while True:
        if keyboard.is_pressed('ctrl+v'):
            time.sleep(0.2)
            if not keyboard.is_pressed('v'):
                break
        time.sleep(0.01)

    # Copy the current cell value to the clipboard
    cell_value = sheet.cell(row=row, column=2).value
    print(cell_value)
    if cell_value is not None:
        pyperclip.copy(str(cell_value))

wb.close()
